"""
All reads/writes to the "database" live here. Two tabs in one Google
Sheet (see gsheet_utils.py):

  Students       — the read-only roster (roll no -> student details).
                   Auto-seeded once from data/students.xlsx, which
                   ships in the repo, so no manual copy-paste is
                   needed on a fresh Google Sheet.
  Registrations  — one row per submitted registration. This is the
                   data that used to disappear on Render restarts;
                   now it lives in the sheet instead of a local file.

This module keeps the exact same public function names/signatures it
had when it was backed by openpyxl + local .xlsx files, so views.py
only needed small changes to the two "download as .xlsx" endpoints.
"""

import io
import os
import threading
from datetime import datetime

import openpyxl

from . import gsheet_utils

STUDENTS_SHEET_NAME = "Students"
REGISTRATIONS_SHEET_NAME = "Registrations"

# Column order matches the original students.xlsx exactly.
STUDENTS_HEADERS = [
    "Slno", "Roll Number", "Student Name", "Father Name",
    "Class Awarded", "CGPA", "Month & Year", "Mobile", "Email",
]
STUDENT_FIELDS = [
    "slno", "roll_no", "name", "father_name", "class_awarded",
    "cgpa", "month_year", "mobile", "email",
]

REGISTRATIONS_HEADERS = [
    "Roll No", "Name", "Attend", "Persons Count",
    "Names", "Contacts", "Relations", "Submitted At",
]

_lock = threading.Lock()

# The students.xlsx that ships in the repo, used only to auto-seed a
# brand-new Google Sheet the first time the Students tab is empty.
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_LOCAL_STUDENTS_SEED_FILE = os.path.join(_BASE_DIR, "data", "students.xlsx")


def _pad(row, width):
    row = list(row) + [""] * (width - len(row))
    return row[:width]


def _seed_students_if_empty(ws):
    if len(ws.get_all_values()) > 1:
        return  # already has data rows
    if not os.path.exists(_LOCAL_STUDENTS_SEED_FILE):
        return
    with _lock:
        if len(ws.get_all_values()) > 1:
            return
        wb = openpyxl.load_workbook(_LOCAL_STUDENTS_SEED_FILE, data_only=True)
        src = wb.active
        rows = []
        for row in src.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            rows.append([
                ("" if c is None else c) for c in _pad(row, len(STUDENTS_HEADERS))
            ])
        if rows:
            ws.append_rows(rows, value_input_option="RAW")


def _students_ws():
    ws = gsheet_utils.get_or_create_worksheet(STUDENTS_SHEET_NAME, STUDENTS_HEADERS)
    _seed_students_if_empty(ws)
    return ws


def _registrations_ws():
    return gsheet_utils.get_or_create_worksheet(REGISTRATIONS_SHEET_NAME, REGISTRATIONS_HEADERS)


def _row_to_student(row):
    return dict(zip(STUDENT_FIELDS, row))


def find_student(roll_no):
    ws = _students_ws()
    target = roll_no.strip().upper()
    for row in ws.get_all_values()[1:]:
        if row and row[1] and str(row[1]).strip().upper() == target:
            return _row_to_student(_pad(row, len(STUDENTS_HEADERS)))
    return None


def get_all_students(search=None):
    ws = _students_ws()
    students = []
    for row in ws.get_all_values()[1:]:
        if not row or not row[1]:
            continue
        student = _row_to_student(_pad(row, len(STUDENTS_HEADERS)))
        if search:
            s = search.strip().lower()
            haystack = f"{student['roll_no']} {student['name']} {student['father_name']}".lower()
            if s not in haystack:
                continue
        students.append(student)
    return students


def is_already_registered(roll_no):
    ws = _registrations_ws()
    target = roll_no.strip().upper()
    for row in ws.get_all_values()[1:]:
        if row and row[0] and str(row[0]).strip().upper() == target:
            return True
    return False


def save_registration(data):
    persons = data.get("persons_list", [])
    names = ", ".join(p.get("name", "") for p in persons)
    contacts = ", ".join(p.get("contact", "") for p in persons)
    relations = ", ".join(p.get("relation", "") for p in persons)

    with _lock:
        ws = _registrations_ws()
        ws.append_row([
            data["roll_no"], data.get("name", ""), data.get("attend", ""), len(persons),
            names, contacts, relations,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ], value_input_option="RAW")


def get_all_registrations():
    ws = _registrations_ws()
    results = []
    for row in ws.get_all_values()[1:]:
        if not row or not row[0]:
            continue
        row = _pad(row, len(REGISTRATIONS_HEADERS))
        names = [n.strip() for n in (row[4] or "").split(",") if n.strip()]
        contacts = [c.strip() for c in (row[5] or "").split(",") if c.strip()]
        relations = [r.strip() for r in (row[6] or "").split(",") if r.strip()]
        persons = list(zip(names, contacts, relations))
        try:
            persons_count = int(row[3]) if row[3] else 0
        except (ValueError, TypeError):
            persons_count = 0
        results.append({
            "roll_no": row[0], "name": row[1], "attend": row[2],
            "persons_count": persons_count, "persons": persons,
            "submitted_at": row[7],
        })
    return results


def export_registrations_xlsx_bytes():
    """In-memory .xlsx snapshot of the Registrations tab, for the
    'download registrations.xlsx' button."""
    wb = openpyxl.Workbook()
    ws_out = wb.active
    ws_out.title = "Registrations"
    ws_out.append(REGISTRATIONS_HEADERS)
    for row in _registrations_ws().get_all_values()[1:]:
        if row and row[0]:
            ws_out.append(_pad(row, len(REGISTRATIONS_HEADERS)))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_students_xlsx_bytes():
    """In-memory .xlsx snapshot of the Students tab, for the
    'download students.xlsx' button."""
    wb = openpyxl.Workbook()
    ws_out = wb.active
    ws_out.title = "Students"
    ws_out.append(STUDENTS_HEADERS)
    for row in _students_ws().get_all_values()[1:]:
        if row and row[1]:
            ws_out.append(_pad(row, len(STUDENTS_HEADERS)))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
